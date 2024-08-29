import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from datetime import datetime
import json
import os


class ExcelCreator:
    def __init__(self, commits_file, analysis_file, keyword_search_file, excel_file, file_path=""):
        self.analysis_file = file_path + analysis_file
        self.excel_file = file_path + excel_file
        self.commits_file = file_path + commits_file
        self.keyword_search_file = file_path + keyword_search_file
        self.sheet_name = os.path.splitext(os.path.basename(analysis_file))[0]  # Sheet named after the file
        self.max_column_width_px = 800  # Maximum column width in pixels for the organization_commits sheet

    def create_workbook(self):
        # Load or create a workbook
        if os.path.exists(self.excel_file):
            workbook = openpyxl.load_workbook(self.excel_file, keep_vba=True)  # Load existing .xlsm with macros
            print(f"Loaded existing Excel file: {self.excel_file}")

            # Check and remove 'Sheet3' if it exists
            if 'Sheet3' in workbook.sheetnames:
                std = workbook['Sheet3']
                workbook.remove(std)
                print("Removed unused Sheet3")
        else:
            workbook = openpyxl.Workbook()
            print(f"Created new Excel file: {self.excel_file}")

        # Step 1: Add commit analysis data
        self.add_commit_analysis_sheet(workbook)

        # Step 2: Add organization commits data
        self.add_organization_commits_sheet(workbook)

        # Step 3: Add keyword search results data
        self.add_keyword_search_results_sheet(workbook)

        # Check and remove 'Sheet3' if it exists
        if 'Sheet3' in workbook.sheetnames:
            std = workbook['Sheet3']
            workbook.remove(std)
            print("Removed unused Sheet3")

        # Save the workbook as a macro-enabled workbook (.xlsm)
        workbook.save(self.excel_file)
        print(f"Excel file '{self.excel_file}' updated successfully.")

    @staticmethod
    def clear_sheet(sheet):
        # Clear the contents of the sheet while keeping the sheet itself
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None

    def add_commit_analysis_sheet(self, workbook):
        # Check if the sheet already exists
        if self.sheet_name in workbook.sheetnames:
            sheet = workbook[self.sheet_name]
            self.clear_sheet(sheet)
            print(f"Cleared existing sheet: {self.sheet_name}")
        else:
            sheet = workbook.create_sheet(self.sheet_name)
            print(f"Created new sheet: {self.sheet_name}")

        # Load data from the commit_analysis.txt file
        data, summary = self.load_commit_analysis()

        # Write data to the worksheet
        self.write_commit_analysis_to_sheet(sheet, data, summary)

    def load_commit_analysis(self):
        # Read the commit_analysis.txt file and parse the data
        data = []
        summary = []
        with open(self.analysis_file, "r") as file:
            lines = file.readlines()

        repo_data = None
        is_summary = False
        for line in lines:
            if line.startswith("Overall Summary:"):
                is_summary = True
                continue

            if is_summary:
                summary.append(line.strip())
            else:
                if line.startswith("Repository:"):
                    if repo_data:
                        data.append(repo_data)
                    repo_name = line.split(":")[1].strip()
                    repo_data = {"repository": repo_name}
                elif line.strip().startswith("- Total commits:"):
                    repo_data["commit_count"] = int(line.split(":")[1].strip())
                elif line.strip().startswith("- Unique committers:"):
                    repo_data["committer_count"] = int(line.split(":")[1].strip())
                elif line.strip().startswith("- Unique authors:"):
                    repo_data["author_count"] = int(line.split(":")[1].strip())

        if repo_data:
            data.append(repo_data)

        return data, summary

    def write_commit_analysis_to_sheet(self, sheet, data, summary):
        # Define headers
        headers = ["Repository", "Total Commits", "Unique Committers", "Unique Authors"]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

        # Write data rows
        for row_num, repo_data in enumerate(data, start=2):
            sheet.cell(row=row_num, column=1, value=repo_data["repository"])
            sheet.cell(row=row_num, column=2, value=repo_data["commit_count"])
            sheet.cell(row=row_num, column=3, value=repo_data["committer_count"])
            sheet.cell(row=row_num, column=4, value=repo_data["author_count"])

            # Apply border and alignment
            for col_num in range(1, 5):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal="center")

        # Write summary in a single row with thick top border
        start_row = len(data) + 3
        summary_data = ["Summary"] + summary[:3]  # Ensure only three summary elements
        for col_num, item in enumerate(summary_data, 1):
            cell = sheet.cell(row=start_row, column=col_num, value=item)
            if col_num == 1:
                cell.border = Border(top=Side(style='thick'))
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        self.adjust_column_width(sheet, max_width_px=500)  # Adjust with a different max width for commit analysis

    @staticmethod
    def adjust_column_width(sheet, max_width_px):
        for col in sheet.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, max_width_px / 7)  # Limit column width
            sheet.column_dimensions[col_letter].width = adjusted_width

    def add_organization_commits_sheet(self, workbook):
        # Create or clear the sheet for organization commits
        sheet_name = "organization_commits"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self.clear_sheet(sheet)
            print(f"Cleared existing sheet: {sheet_name}")
        else:
            sheet = workbook.create_sheet(sheet_name)
            print(f"Created new sheet: {sheet_name}")

        # Load the organization commits data
        with open(self.commits_file, "r") as file:
            commits_data = json.load(file)

        # Define headers
        headers = ["Repository", "Message", "Committer", "Author", "SHA", "Date"]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data rows
        row_num = 2
        for repo, commits in commits_data.items():
            for commit in commits:
                if commit:
                    # Extract fields safely
                    commit_message = commit.get("commit", {}).get("message", "unknown")
                    committer_info = commit.get("commit", {}).get("committer", {})
                    committer = committer_info.get("name", "unknown") if committer_info else "unknown"
                    author_info = commit.get("commit", {}).get("author", {})
                    author = author_info.get("name", "unknown") if author_info else "unknown"
                    commit_date = committer_info.get("date", "unknown") if committer_info else "unknown"
                    commit_date = commit_date.replace("T", " ").replace("Z", "")
                    sha = commit.get("sha", "unknown")

                    # Write values to the sheet
                    sheet.cell(row=row_num, column=1, value=repo)
                    sheet.cell(row=row_num, column=2, value=commit_message)
                    sheet.cell(row=row_num, column=3, value=committer)
                    sheet.cell(row=row_num, column=4, value=author)
                    sheet.cell(row=row_num, column=5, value=sha)
                    sheet.cell(row=row_num, column=6, value=commit_date)

                    # Enable word wrapping for the message column
                    sheet.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True)

                    row_num += 1

        # Adjust column widths with word wrapping
        self.adjust_column_width(sheet, max_width_px=self.max_column_width_px)

    def add_keyword_search_results_sheet(self, workbook):
        # Create or clear the sheet for keyword search results
        sheet_name = "keyword_search_results"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self.clear_sheet(sheet)
            print(f"Cleared existing sheet: {sheet_name}")
        else:
            sheet = workbook.create_sheet(sheet_name)
            print(f"Created new sheet: {sheet_name}")

        # Load the keyword search results data
        keyword_search_file = self.keyword_search_file
        with open(keyword_search_file, "r") as file:
            search_results_data = json.load(file)

        # Define headers
        headers = [
            "Commit Date", "Repository", "Message", "Unique Finds", "Total Instances", "Committer", "Author", "SHA"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data rows
        row_num = 2
        for sha, data in search_results_data.items():
            repo = data.get("repo", "unknown")
            commit_data = data.get("commit", {})

            # Extract commit information
            message = commit_data.get("message", "unknown")
            unique_finds = commit_data.get("unique_finds", 0)
            total_instances = commit_data.get("total_instances", 0)
            committer = commit_data.get("committer", "unknown")
            author = commit_data.get("author", "unknown")
            commit_date = commit_data.get("commit_date", "unknown")
            commit_date.replace("T", " ").replace("Z", "")

            # Write values to the sheet
            sheet.cell(row=row_num, column=1, value=commit_date)
            sheet.cell(row=row_num, column=2, value=repo)
            sheet.cell(row=row_num, column=3, value=message)
            sheet.cell(row=row_num, column=4, value=unique_finds)
            sheet.cell(row=row_num, column=5, value=total_instances)
            sheet.cell(row=row_num, column=6, value=committer)
            sheet.cell(row=row_num, column=7, value=author)
            sheet.cell(row=row_num, column=8, value=sha)

            # Enable word wrapping for the message column
            sheet.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True)

            row_num += 1

        # Adjust column widths with word wrapping
        self.adjust_column_width(sheet, max_width_px=self.max_column_width_px)
